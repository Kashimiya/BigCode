from openpyxl import Workbook
import os

class TeacherHuangInfluence:
    def teacherHuangInfluence(self):

        excelPath=os.path.abspath('../..')+'\\doc\\DataAnalysis.xlsx'
        workbook=Workbook()
        sheet=workbook.active
        sheet.title='title'
        # lower和 higher 代表输入做题数目区间的上限和下限
        path = os.path.abspath('../..') + '\\doc\\Result'
        doc = open(path, 'r')
        for i in range(0, 121):
            temp = doc.readline()
            temp = temp.strip('\n')
            if len(temp) == 0:
                continue
            if i == 50:
                # 找到黄老师催促以前的数据
                hlBefore = temp
                hlBefore = hlBefore.split(',')
                hlBefore.remove(hlBefore[0])
                for j in range(0, len(hlBefore)):
                    hlBefore[j] = int(hlBefore[j])
            elif i == 56:
                # 找到黄老师催促以后的数据
                hlAfter = temp
                hlAfter = hlAfter.split(',')
                hlAfter.remove(hlAfter[0])
                for j in range(0, len(hlAfter)):
                    hlAfter[j] = int(hlAfter[j])
            elif i == 118:
                # 最后ddl的数据
                endTime = temp
                endTime = endTime.split(',')
                endTime.remove(endTime[0])
                # 将最后都没有做超过20题的人剔除，默认退课
                dropOut = []
                for j in range(0, len(endTime)):
                    endTime[j] = int(endTime[j])
                    if endTime[j] < 20:
                        dropOut.append(j)
        # distribution是每个同学平均每日增长解题数
        count=1
        for i in range(0, len(hlBefore)):
            if i not in dropOut:
                distribution=((hlAfter[i] - hlBefore[i]) / 3 - hlBefore[i] / 10)
                sheet.cell(count,1).value=hlBefore[i]
                sheet.cell(count,2).value=distribution
                count=count+1

        workbook.save(excelPath)

if __name__ == '__main__':
    a=TeacherHuangInfluence()
    a.teacherHuangInfluence()
